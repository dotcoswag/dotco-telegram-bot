"""
export.py
---------
Converts a scraper CSV output to a formatted XLSX file.
Applies column widths, header styling, and lead score color coding.

Usage (from main.py or standalone):
    python export.py resultados/dotco_leads_YYYYMMDD_HHMMSS.csv
"""

import csv
import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Install openpyxl: pip install openpyxl")
    sys.exit(1)


# Column widths (col_name: width)
COL_WIDTHS = {
    "business_id": 24,
    "nombre": 32,
    "tipo": 22,
    "subtipo": 28,
    "rating": 8,
    "review_count": 12,
    "verified": 10,
    "business_status": 16,
    "lead_score": 10,
    "photos_count": 14,
    "direccion": 36,
    "city": 18,
    "state": 14,
    "district": 16,
    "latitude": 12,
    "longitude": 12,
    "telefono": 16,
    "email": 28,
    "website": 32,
    "instagram": 28,
    "facebook": 28,
    "linkedin": 28,
    "twitter": 22,
    "youtube": 28,
    "emails_extra": 32,
    "link_google_maps": 36,
    "booking_link": 28,
    "menu_link": 28,
    "order_link": 28,
    "localidad_buscada": 20,
    "categoria_buscada": 22,
    "provincia": 16,
}

# Lead score colors (score → hex fill)
SCORE_COLORS = {
    5: "1A7F4B",  # Dark green — hot lead
    4: "52B788",  # Green
    3: "F4A261",  # Orange
    2: "E76F51",  # Red-orange
    1: "C1121F",  # Red
    0: "AAAAAA",  # Gray
}

HEADER_FILL = PatternFill("solid", fgColor="1B1B2F")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
BORDER_SIDE = Side(style="thin", color="DDDDDD")
THIN_BORDER = Border(
    left=BORDER_SIDE, right=BORDER_SIDE,
    top=BORDER_SIDE, bottom=BORDER_SIDE
)


def csv_to_xlsx(csv_path: str, xlsx_path: str = None) -> str:
    if not xlsx_path:
        xlsx_path = csv_path.replace(".csv", ".xlsx")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DotCo Leads"

    # Freeze header row
    ws.freeze_panes = "A2"

    headers = None

    with open(csv_path, encoding="utf-8") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []

        # Write header row
        for col_idx, col_name in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name.upper().replace("_", " "))
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

        # Write data rows
        for row_idx, row in enumerate(reader, 2):
            for col_idx, col_name in enumerate(headers, 1):
                value = row.get(col_name, "")

                # Convert numeric fields
                if col_name in ("rating", "latitude", "longitude"):
                    try:
                        value = float(value)
                    except (ValueError, TypeError):
                        pass
                elif col_name in ("review_count", "lead_score", "photos_count"):
                    try:
                        value = int(value)
                    except (ValueError, TypeError):
                        pass
                elif col_name == "verified":
                    value = "Yes" if str(value).lower() in ("true", "1", "yes") else ""

                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=False)

                # Color-code lead_score column
                if col_name == "lead_score":
                    try:
                        score = int(value)
                        hex_color = SCORE_COLORS.get(score, "FFFFFF")
                        cell.fill = PatternFill("solid", fgColor=hex_color)
                        cell.font = Font(color="FFFFFF", bold=True)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    except (ValueError, TypeError):
                        pass

                # Clickable links
                if col_name in ("website", "link_google_maps", "instagram", "facebook") and value and str(value).startswith("http"):
                    cell.hyperlink = str(value)
                    cell.font = Font(color="1155CC", underline="single")

    # Set column widths
    if headers:
        for col_idx, col_name in enumerate(headers, 1):
            width = COL_WIDTHS.get(col_name, 18)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Auto-filter on header row
    ws.auto_filter.ref = ws.dimensions

    # Row height for data rows
    for row in ws.iter_rows(min_row=2):
        ws.row_dimensions[row[0].row].height = 16

    # Header row height
    ws.row_dimensions[1].height = 22

    wb.save(xlsx_path)
    return xlsx_path


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python export.py <path_to_csv>")
        sys.exit(1)

    csv_file = sys.argv[1]
    if not os.path.exists(csv_file):
        print(f"File not found: {csv_file}")
        sys.exit(1)

    out = csv_to_xlsx(csv_file)
    print(f"✅ Exported: {out}")
